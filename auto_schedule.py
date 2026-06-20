import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color
import re
from datetime import date, timedelta, datetime
import calendar
from pathlib import Path

PROJECT_DIR = Path(r"H:\Gemini\pss")

def find_template_11404() -> Path:
    candidates = [p for p in PROJECT_DIR.glob("*.xlsx") if "11404" in p.stem and not p.name.startswith("~$") and not p.name.startswith("._")]
    if not candidates: raise FileNotFoundError("找不到 11404 樣板檔案。")
    return candidates[0]

def run_scheduler(filepath: str):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    fonts = {
        '周': Font(name='標楷體', size=12, bold=True, color='FF0070C0'),
        '鄭': Font(name='標楷體', size=12, bold=True, color='FFFF0000'),
        '石': Font(name='標楷體', size=12, bold=True, color='FF007A37'),
    }
    align = Alignment(horizontal='center', vertical='center')

    def parse_dates(cell_val):
        if not cell_val: return []
        return [int(x) for x in re.findall(r'\d+', str(cell_val))]

    offs = {'周': [], '石': [], '鄭': []}
    doc_rows = {'周': 4, '石': 5, '鄭': 6}
    
    for doc, r in doc_rows.items():
        offs[doc] = parse_dates(ws.cell(row=r, column=12).value)
        
    yushan_dates = parse_dates(ws.cell(row=7, column=13).value)
    
    reqs = {}
    for doc, r in doc_rows.items():
        reqs[doc] = {
            '早': parse_dates(ws.cell(row=r, column=13).value),
            '午': parse_dates(ws.cell(row=r, column=14).value),
            '午夜': parse_dates(ws.cell(row=r, column=15).value),
            '夜': parse_dates(ws.cell(row=r, column=16).value),
        }
        
    date_cells = {}
    day_to_weekday = {}
    for r in [3, 8, 13, 18, 23]:
        for c in range(2, 9):
            val = ws.cell(row=r, column=c).value
            if val:
                if isinstance(val, datetime): day = val.day
                else:
                    match = re.search(r'(\d+)日', str(val))
                    if match: day = int(match.group(1))
                    else:
                        match = re.search(r'\d+/(\d+)', str(val))
                        if match: day = int(match.group(1))
                        else: continue
                date_cells[day] = (r, c)
                day_to_weekday[day] = c - 2 

    # --- P1 ---
    grid_rows = [4,5,6, 9,10,11, 14,15,16, 19,20,21, 24,25,26]
    for r in grid_rows:
        for c in range(2, 9):
            ws.cell(row=r, column=c).value = None

    def apply_shift(doc, day, shift_offset):
        if day in date_cells:
            r, c = date_cells[day]
            cell = ws.cell(row=r + shift_offset, column=c)
            cell.value = doc
            cell.font = fonts[doc]
            cell.alignment = align

    for doc, shifts in reqs.items():
        for day in shifts['早']: apply_shift(doc, day, 1)
        for day in shifts['午']: apply_shift(doc, day, 2)
        for day in shifts['夜']: apply_shift(doc, day, 3)
        for day in shifts['午夜']: 
            apply_shift(doc, day, 2)
            apply_shift(doc, day, 3)
            
    print("\n>> P1 (指定需求) 完成。")

    day_shifts = {day: {1: None, 2: None, 3: None} for day in date_cells}
    total_shifts = {'周': 0, '鄭': 0, '石': 0}
    
    for day, (r, c) in sorted(date_cells.items()):
        for offset in [1, 2, 3]:
            doc = ws.cell(row=r + offset, column=c).value
            if doc in ['周', '鄭', '石']:
                day_shifts[day][offset] = doc
                total_shifts[doc] += 1
                
    def get_doc_shifts_today(doc, day):
        return [s for s, d in day_shifts[day].items() if d == doc]
        
    def can_work(doc, day, shift):
        weekday = day_to_weekday[day]
        if day in offs[doc]: return False
        if doc == '周' and weekday in (5, 6) and shift == 1 and day in yushan_dates: return False
            
        my_shifts = get_doc_shifts_today(doc, day)
        if doc == '石':
            if weekday == 5 and shift in (2, 3): return False
            if shift == 1 and (2 in my_shifts or 3 in my_shifts): return False
            if shift in (2, 3) and 1 in my_shifts: return False
        if doc == '鄭':
            if shift == 1 and (2 in my_shifts or 3 in my_shifts): return False
            if shift in (2, 3) and 1 in my_shifts: return False
            if weekday in (0, 2, 4):
                if shift == 2 and 3 in my_shifts: return False
                if shift == 3 and 2 in my_shifts: return False
        return True

    # --- P2 & P3 ---
    prototypes = {
        0: {1: '周', 2: '鄭', 3: '石'}, 
        1: {1: '周', 2: '石', 3: '鄭'}, 
        2: {1: '鄭', 2: '石', 3: '周'}, 
        3: {1: '周', 2: '鄭', 3: '石'}, 
        4: {1: '周', 2: '石', 3: '鄭'}, 
    }
    
    def assign(day, shift, doc):
        r, c = date_cells[day]
        cell = ws.cell(row=r + shift, column=c)
        cell.value = doc
        cell.font = fonts[doc]
        cell.alignment = align
        day_shifts[day][shift] = doc
        total_shifts[doc] += 1
        
    for day in sorted(date_cells.keys()):
        weekday = day_to_weekday[day]
        if weekday in prototypes:
            proto_day = prototypes[weekday]
            for shift in [1, 2, 3]:
                doc = proto_day[shift]
                if not day_shifts[day][shift]:
                    if day not in offs[doc] and len(get_doc_shifts_today(doc, day)) == 0:
                        assign(day, shift, doc)
        elif weekday == 5:
            doc = '石'
            if not day_shifts[day][1] and day not in offs[doc] and len(get_doc_shifts_today(doc, day)) == 0:
                assign(day, 1, doc)
                
    print(">> P2 & P3 (固定原型/特殊死律) 完成。")

    # --- P4 ---
    empty_cells = []
    for day in sorted(date_cells.keys()):
        for shift in [1, 2, 3]:
            if not day_shifts[day][shift]:
                empty_cells.append((day, shift))
                
    for day, shift in empty_cells:
        valid_cands = [d for d in ['周', '鄭', '石'] if can_work(d, day, shift)]
        shift_name = {1:'早', 2:'午', 3:'晚'}[shift]
        
        if not valid_cands:
            print(f"\n[⚠️ 衝突警告] 發現無法排班的空缺：{day}日 {shift_name}診！")
            print("  原因：所有醫師均受限或休假。")
            while True:
                ans = input("  請問要強制指派給誰？ (1=周, 2=鄭, 3=石, 0=先留空): ").strip()
                doc_map = {'1':'周', '2':'鄭', '3':'石'}
                if ans == '0': break
                elif ans in doc_map:
                    assign(day, shift, doc_map[ans])
                    break
            continue
            
        def score(doc):
            return (len(get_doc_shifts_today(doc, day)), total_shifts[doc])
            
        best_doc = sorted(valid_cands, key=score)[0]
        assign(day, shift, best_doc)
        
    print("\n>> P4 (缺額補強) 完成。")

    # --- Calc Stats & Final Balance ---
    def update_excel_stats():
        ws.cell(row=10, column=13).value = total_shifts['周']
        ws.cell(row=11, column=13).value = total_shifts['鄭']
        ws.cell(row=12, column=13).value = total_shifts['石']
        for doc, r in [('周', 4), ('石', 5), ('鄭', 6)]:
            two_shift_days = sum(1 for day in date_cells if len(get_doc_shifts_today(doc, day)) >= 2)
            work_days = sum(1 for day in date_cells if len(get_doc_shifts_today(doc, day)) > 0)
            ws.cell(row=r, column=18).value = two_shift_days
            if doc == '周': ws.cell(row=10, column=16).value = work_days
            elif doc == '鄭': ws.cell(row=11, column=16).value = work_days
            elif doc == '石': ws.cell(row=12, column=16).value = work_days

    def safe_save():
        while True:
            try:
                wb.save(filepath)
                break
            except PermissionError:
                print(f"\n[❌ 錯誤] 檔案儲存失敗！")
                input(f"請先在 Excel 中關閉 {Path(filepath).name}，關閉後按 Enter 鍵重試...")

    update_excel_stats()
    safe_save()
    print(f"\n✅ 初步排班結果已儲存至：{filepath}！")
    print("👉 您現在可以打開 Excel 查看排班結果。")
            
    while True:
        print("\n=== 🎯 目前總結算 ===")
        print(f"周: {total_shifts['周']}, 鄭: {total_shifts['鄭']}, 石: {total_shifts['石']}")
        if total_shifts['周'] == total_shifts['鄭'] == total_shifts['石']:
            print("✨ 完美平衡！(31:31:31) ✨")
            break
            
        ans = input("\n診次不平衡，請問是否要手動換班？(y/n): ").strip().lower()
        if ans != 'y': break
        
        day_str = input("請輸入要修改的日期 (例如 10): ").strip()
        if not day_str.isdigit() or int(day_str) not in date_cells:
            print("日期無效。")
            continue
            
        shift_str = input("請輸入班別 (1=早, 2=午, 3=晚): ").strip()
        if shift_str not in ['1','2','3']:
            print("班別無效。")
            continue
            
        doc_str = input("請輸入要換成哪位醫師 (1=周, 2=鄭, 3=石, 0=清除): ").strip()
        if doc_str not in ['0','1','2','3']:
            print("醫師代碼無效。")
            continue
            
        doc_map = {'1':'周', '2':'鄭', '3':'石'}
        day = int(day_str)
        shift = int(shift_str)
        
        old_doc = day_shifts[day][shift]
        if old_doc: total_shifts[old_doc] -= 1
            
        if doc_str == '0':
            r, c = date_cells[day]
            ws.cell(row=r+shift, column=c).value = None
            day_shifts[day][shift] = None
        else:
            assign(day, shift, doc_map[doc_str])
            
        update_excel_stats()
        safe_save()
        print("✅ 修改已更新並儲存！")
            
    print(f"\n✅ 排班結束！檔案已確認儲存。")

def main():
    print("================================")
    print("   賜安診所自動排班系統 v1.0    ")
    print("================================")
    while True:
        ym = input("\n請輸入要排班的年月 (例如 11506) 或按 Q 離開: ").strip()
        if ym.lower() == 'q': return
        if len(ym) >= 4 and ym.isdigit():
            break
        print("格式錯誤，請重試。")
        
    roc_year = int(ym[:-2])
    month = int(ym[-2:])
    western_year = roc_year + 1911
    
    output_path = PROJECT_DIR / f"賜安診所{roc_year}{month:02d}班表_空白.xlsx"
    
    ans = input(f"\n是否需要自動產生空白班表 {output_path.name} ? (y/n): ").strip().lower()
    if ans == 'y':
        template = find_template_11404()
        wb = openpyxl.load_workbook(template)
        ws = wb.active
        
        header_rows = [2, 7, 12, 17, 22]
        date_rows = [3, 8, 13, 18, 23]
        weekday_labels = [ws.cell(row=7, column=c).value for c in range(2, 9)]
        shift_rows = []
        for r in date_rows: shift_rows.extend([r + 1, r + 2, r + 3])
        
        ws["A1"] = f"賜安診所{roc_year}年{month}月班表"
        
        for r in date_rows:
            for c in range(2, 9): ws.cell(row=r, column=c).value = None
        for r in header_rows:
            for c in range(2, 9): ws.cell(row=r, column=c).value = None
        for r in shift_rows:
            for c in range(2, 9): ws.cell(row=r, column=c).value = None
            
        current = date(western_year, month, 1)
        week_index = 0
        days_in_month = calendar.monthrange(western_year, month)[1]
        
        for _ in range(days_in_month):
            if week_index >= 5:
                week_index = 0  # 解決 6 週月份（如8月）的問題，將最後幾天繞回第一週的空格
            row = date_rows[week_index]
            col = current.weekday() + 2
            ws.cell(row=row, column=col).value = f"{month}月{current.day}日"
            if current.weekday() == 6: week_index += 1
            current += timedelta(days=1)
            
        for h_row, d_row in zip(header_rows, date_rows):
            for idx, c in enumerate(range(2, 9)):
                if ws.cell(row=d_row, column=c).value:
                    ws.cell(row=h_row, column=c).value = weekday_labels[idx]
                    
        # Apply seasonal color pattern
        if 1 <= month <= 3:
            season_fill = PatternFill(patternType='solid', fgColor='FF7EE7EA')
        elif 4 <= month <= 6:
            season_fill = PatternFill(patternType='solid', fgColor='FFFFC000')
        elif 7 <= month <= 9:
            season_fill = PatternFill(patternType='solid', fgColor='FF92D050')
        else:
            season_fill = PatternFill(patternType='solid', fgColor=Color(theme=7))
            
        for r in header_rows + date_rows:
            ws.cell(row=r, column=1).fill = season_fill  # 第一欄的「第幾週」底色同步
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = season_fill
                    
        for r in range(2, 27):
            for c in range(11, 21):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is None: continue
                text = str(v).strip()
                keep_tokens = {"off", "OFF", "周", "鄭", "石", "俞珊", "早", "午", "夜", "午夜", "總診次", "總天數", "兩班", "週六早上 石", "2 4不可連診"}
                if text in keep_tokens: continue
                if any(ch.isdigit() for ch in text): cell.value = None
                
        wb.save(output_path)
        print(f"\n✅ 空白班表已產生：{output_path.name}")
        
    print(f"\n👉 請打開 {output_path.name}，")
    print("在 K~P 欄填入各醫師的「休假」與「指定班次」。")
    input("填好並「存檔」後，請回到這裡按 Enter 鍵繼續排班...")
    
    run_scheduler(str(output_path))

if __name__ == "__main__":
    main()
